"""
results_store.py — Gönderim sonuçlarını RAM'de tut ve görev bitiminde Excel'e yaz.
- Broadcast sayfası: phone, name, Status, Timestamp, FinalMessage, Error
- Summary sayfası: toplam, sent, failed, success_rate (%)
- Satır renkleme: koşullu biçimlendirme (Status='Sent' yeşil, 'Failed' kırmızı)
"""
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
import os

GREEN = "C6EFCE"
RED   = "FFC7CE"

class ResultAggregator:
    """Tüm sonuçları RAM'de tutar; sonunda Excel'e yazar."""
    def __init__(self):
        self.rows: List[Dict[str, str]] = []

    def add(self, phone: str, name: str, status: str, timestamp: str, final_message: str, error: str = ""):
        self.rows.append({
            "phone": phone,
            "name": name,
            "Status": status,              # 'Sent' | 'Failed'
            "Timestamp": timestamp,
            "FinalMessage": final_message,
            "Error": error
        })

    def _write_broadcast_sheet(self, wb: Workbook):
        ws = wb.create_sheet(title="Broadcast")
        headers = ["phone", "name", "Status", "Timestamp", "FinalMessage", "Error"]
        ws.append(headers)
        for row in self.rows:
            ws.append([row.get(h, "") for h in headers])

        # Koşullu biçimlendirme: tüm satırı renkle
        status_col = headers.index("Status") + 1
        max_col_letter = get_column_letter(ws.max_column)
        rng = f"A2:{max_col_letter}{ws.max_row}"
        status_letter = get_column_letter(status_col)
        formula_sent = f'INDIRECT("{status_letter}"&ROW())="Sent"'
        formula_failed = f'INDIRECT("{status_letter}"&ROW())="Failed"'
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[formula_sent],
                        fill=PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid"))
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[formula_failed],
                        fill=PatternFill(start_color=RED, end_color=RED, fill_type="solid"))
        )

        # Kolon genişlikleri (hafif optimizasyon)
        for col_idx, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(60, len(h) + 2))

    def _write_summary_sheet(self, wb: Workbook):
        ws = wb.create_sheet(title="Summary")
        total = len(self.rows)
        sent = sum(1 for r in self.rows if r["Status"] == "Sent")
        failed = total - sent
        success_rate = (sent / total * 100.0) if total else 0.0

        ws.append(["Metric", "Value"])
        ws.append(["Total", total])
        ws.append(["Sent", sent])
        ws.append(["Failed", failed])
        ws.append(["SuccessRate(%)", round(success_rate, 2)])

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18

    def export_excel(self, out_path: str):
        """
        RAM'deki verileri tek seferde yeni bir Excel dosyasına yazar.
        Not: openpyxl yeni kitapta varsayılan 'Sheet' oluşturur; onu silip kendi sayfalarımızı ekliyoruz.
        """
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        wb = Workbook()
        # Varsayılan boş sayfayı kaldır
        default = wb.active
        wb.remove(default)

        self._write_broadcast_sheet(wb)
        self._write_summary_sheet(wb)

        wb.save(out_path)
        wb.close()
